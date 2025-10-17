#!/usr/bin/env python3
"""
Evaluation Data Analyzer for WL EVO AVLDrive Heatmap Tool

This script analyzes evaluation metric data pasted by users into the Evaluation Sheet.
It provides statistical analysis, data quality checks, and insights about:
- P1 status distribution (Red/Yellow/Green)
- Drivability and Dynamism ratings
- Overall performance metrics
- Data completeness and quality
"""

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from collections import Counter
import json

try:
    import openpyxl
    from openpyxl.styles import Color
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


class EvaluationDataAnalyzer:
    """Analyzes evaluation data from the Excel workbook"""
    
    # P1 color mappings (RGB values)
    P1_COLORS = {
        'RED': (225, 0, 0),
        'GREEN': (0, 176, 80),
        'YELLOW': (227, 225, 0)
    }
    
    def __init__(self, excel_path: str):
        """Initialize the analyzer with the Excel file path"""
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        self.eval_sheet = None
        self.header_row = None
        self.use_case_col = None
        self.data_rows = []
        
    def find_evaluation_sheet(self) -> bool:
        """Find and load the Evaluation Sheet"""
        for sheet_name in self.workbook.sheetnames:
            if 'evaluation' in sheet_name.lower():
                self.eval_sheet = self.workbook[sheet_name]
                return True
        return False
    
    def detect_header(self) -> bool:
        """Detect the header row containing 'USE CASE'"""
        if not self.eval_sheet:
            return False
        
        # Search first 20 rows for header
        for row_idx in range(1, 21):
            for col_idx in range(1, 201):
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and 'USE CASE' in str(cell_value).upper():
                    self.header_row = row_idx
                    self.use_case_col = col_idx
                    return True
        return False
    
    def find_column_by_header(self, header_text: str, start_col: int = 1, end_col: int = 200) -> Optional[int]:
        """Find a column by searching for header text"""
        if not self.header_row:
            return None
        
        # Check a band around the header row (merged cells)
        for col_idx in range(start_col, end_col + 1):
            for row_offset in range(-1, 3):
                row_idx = self.header_row + row_offset
                if row_idx < 1:
                    continue
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and header_text.upper() in str(cell_value).upper():
                    return col_idx
        return None
    
    def extract_data_rows(self) -> List[Dict]:
        """Extract all data rows from the evaluation sheet"""
        if not self.header_row or not self.use_case_col:
            return []
        
        data_rows = []
        first_data_row = self.header_row + 1
        
        # Find columns
        driv_col = self.find_column_by_header("DRIVABILITY")
        dyn_col = self.find_column_by_header("DYNAMISM")
        
        if not driv_col:
            print("Warning: Could not locate Drivability column")
            # Default based on typical structure
            driv_col = 6
        
        if not dyn_col:
            print("Warning: Could not locate Dynamism column, using typical structure")
            # Dynamism typically starts much later, skip for now
            dyn_col = driv_col + 20
        
        # Find P1 columns - typically at specific positions
        # Based on the structure: Drivability P1 is usually at column F (6)
        driv_p1_col = self._find_p1_column_simple(driv_col, "Current Status")
        
        # Find vehicle data columns (user-pasted data)
        vehicle_cols = self._find_vehicle_data_columns()
        
        # The actual use case data is in column C (merged with A)
        data_col = 3
        
        # Extract data rows
        for row_idx in range(first_data_row, first_data_row + 500):
            use_case = self.eval_sheet.cell(row_idx, data_col).value
            
            # Skip empty rows and category headers
            if not use_case or str(use_case).strip() == "":
                continue
            
            use_case_str = str(use_case).strip()
            
            # Skip rows that are just dots or category headers
            if use_case_str in ["●", "Drive away", "Accelerations", "Decelerations", 
                               "Constant Speeds", "Gearbox behaviour", "Constant speeds",
                               "Status", "Drivability Lowest Events", "Dynamism Lowest Events"]:
                continue
            
            # Get P1 status from drivability column
            driv_p1_status = self._get_p1_status(row_idx, driv_p1_col) if driv_p1_col else "Unspecified"
            
            # Get vehicle data values
            vehicle_data = {}
            for col_idx, vehicle_name in vehicle_cols.items():
                val = self.eval_sheet.cell(row_idx, col_idx).value
                if val is not None and str(val).strip() not in ["", "●"]:
                    vehicle_data[vehicle_name] = val
            
            row_data = {
                'row_number': row_idx,
                'use_case': use_case_str,
                'drivability_p1': driv_p1_status,
                'vehicle_data': vehicle_data,
                'has_vehicle_data': len(vehicle_data) > 0
            }
            data_rows.append(row_data)
        
        self.data_rows = data_rows
        return data_rows
    
    def _find_p1_column_simple(self, start_col: int, status_text: str) -> Optional[int]:
        """Find P1 column by looking for the pattern"""
        if not self.header_row:
            return None
        
        # Look for "Current Status" in row header_row - 1, then P1 in header_row
        for col_idx in range(start_col, start_col + 10):
            # Check if this column or nearby has "Current Status"
            for row_offset in range(-1, 2):
                row_idx = self.header_row + row_offset
                if row_idx < 1:
                    continue
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and status_text.upper() in str(cell_value).upper():
                    # Found Current Status, now look for P1 in the header row
                    p1_val = self.eval_sheet.cell(self.header_row, col_idx).value
                    if p1_val and str(p1_val).strip().upper() == 'P1':
                        return col_idx
                    # Try next column
                    for offset in range(0, 3):
                        p1_val = self.eval_sheet.cell(self.header_row, col_idx + offset).value
                        if p1_val and str(p1_val).strip().upper() == 'P1':
                            return col_idx + offset
        
        return None
    
    def _find_vehicle_data_columns(self) -> Dict[int, str]:
        """Find columns containing vehicle test data (user-pasted)"""
        vehicle_cols = {}
        
        if not self.header_row:
            return vehicle_cols
        
        # Look for vehicle names in row header_row - 1
        # Typically starts around column L (12) or later
        for col_idx in range(10, self.eval_sheet.max_column + 1):
            # Check row before header for vehicle names
            vehicle_name = self.eval_sheet.cell(self.header_row - 1, col_idx).value
            if vehicle_name and str(vehicle_name).strip() not in ["", "●", "P1", "P2", "P3"]:
                vehicle_name_str = str(vehicle_name).strip()
                # Filter out typical header values
                if not any(x in vehicle_name_str.upper() for x in ["CURRENT", "STATUS", "SOPM", "PREDICTION", "DRIVABILITY", "DYNAMISM"]):
                    vehicle_cols[col_idx] = vehicle_name_str
        
        return vehicle_cols
    
    def _find_p1_column(self, start_col: int, end_col: int) -> Optional[int]:
        """Find P1 column within a range"""
        if not self.header_row:
            return None
        
        # First find "Current Status"
        curr_status_col = None
        for col_idx in range(start_col, end_col + 1):
            for row_offset in range(-1, 3):
                row_idx = self.header_row + row_offset
                if row_idx < 1:
                    continue
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and 'CURRENT STATUS' in str(cell_value).upper():
                    curr_status_col = col_idx
                    break
            if curr_status_col:
                break
        
        if not curr_status_col:
            return None
        
        # Find P1 column after Current Status
        for col_idx in range(curr_status_col, end_col + 1):
            for row_offset in range(-1, 3):
                row_idx = self.header_row + row_offset
                if row_idx < 1:
                    continue
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and str(cell_value).strip().upper() == 'P1':
                    return col_idx
        
        return None
    
    def _find_rightmost_column(self, header_text: str) -> Optional[int]:
        """Find the rightmost column with the given header text"""
        if not self.header_row:
            return None
        
        rightmost_col = None
        max_col = self.eval_sheet.max_column
        
        for col_idx in range(max_col, 0, -1):
            for row_offset in range(-1, 3):
                row_idx = self.header_row + row_offset
                if row_idx < 1:
                    continue
                cell_value = self.eval_sheet.cell(row_idx, col_idx).value
                if cell_value and str(cell_value).strip().upper() == header_text.upper():
                    return col_idx
        
        return rightmost_col
    
    def _get_p1_status(self, row_idx: int, col_idx: int) -> str:
        """Get P1 status (Red/Yellow/Green) from cell color or value"""
        if not col_idx:
            return "Unspecified"
        
        cell = self.eval_sheet.cell(row_idx, col_idx)
        
        # Check cell color
        if cell.fill and cell.fill.fgColor:
            color = cell.fill.fgColor
            rgb_str = None
            
            # Handle different color types
            if hasattr(color, 'rgb') and color.rgb:
                if isinstance(color.rgb, str):
                    rgb_str = color.rgb
                elif hasattr(color.rgb, 'index'):
                    # Indexed color - skip for now
                    pass
            
            if rgb_str and isinstance(rgb_str, str):
                if len(rgb_str) == 8:  # ARGB format
                    rgb_str = rgb_str[2:]  # Remove alpha channel
                
                try:
                    r = int(rgb_str[0:2], 16)
                    g = int(rgb_str[2:4], 16)
                    b = int(rgb_str[4:6], 16)
                    
                    # Match against known P1 colors with tolerance
                    if abs(r - 225) < 30 and g < 50 and b < 50:
                        return "Red"
                    elif r < 50 and abs(g - 176) < 30 and abs(b - 80) < 30:
                        return "Green"
                    elif abs(r - 227) < 30 and abs(g - 225) < 30 and b < 50:
                        return "Yellow"
                except (ValueError, IndexError):
                    pass
        
        # Check cell value
        cell_value = str(cell.value).upper().strip() if cell.value else ""
        if cell_value in ['GREEN', 'G']:
            return "Green"
        elif cell_value in ['YELLOW', 'Y']:
            return "Yellow"
        elif cell_value in ['RED', 'R']:
            return "Red"
        
        return "Unspecified"
    
    def analyze_data(self) -> Dict:
        """Perform comprehensive analysis on the data"""
        if not self.data_rows:
            return {}
        
        analysis = {
            'total_rows': len(self.data_rows),
            'use_cases': [],
            'drivability_p1_distribution': {},
            'vehicle_data_summary': {},
            'per_vehicle_statistics': {},
            'data_quality': {},
            'insights': []
        }
        
        # Count distributions
        driv_p1_counter = Counter()
        vehicle_data_rows = 0
        all_vehicles = set()
        vehicle_stats = {}
        
        for row in self.data_rows:
            analysis['use_cases'].append(row['use_case'])
            
            # Drivability P1
            driv_status = row.get('drivability_p1', 'Unspecified') or "Unspecified"
            driv_p1_counter[driv_status] += 1
            
            # Vehicle data
            if row.get('has_vehicle_data', False):
                vehicle_data_rows += 1
                for vehicle, value in row.get('vehicle_data', {}).items():
                    all_vehicles.add(vehicle)
                    
                    # Initialize vehicle stats
                    if vehicle not in vehicle_stats:
                        vehicle_stats[vehicle] = {
                            'count': 0,
                            'sum': 0,
                            'min': None,
                            'max': None,
                            'values': []
                        }
                    
                    # Update stats if value is numeric
                    if isinstance(value, (int, float)):
                        stats = vehicle_stats[vehicle]
                        stats['count'] += 1
                        stats['sum'] += value
                        stats['values'].append(value)
                        
                        if stats['min'] is None or value < stats['min']:
                            stats['min'] = value
                        if stats['max'] is None or value > stats['max']:
                            stats['max'] = value
        
        # Calculate averages
        # Calculate per-vehicle averages and statistics
        per_vehicle_stats = {}
        for vehicle, stats in vehicle_stats.items():
            if stats['count'] > 0:
                per_vehicle_stats[vehicle] = {
                    'data_points': stats['count'],
                    'average': stats['sum'] / stats['count'],
                    'min': stats['min'],
                    'max': stats['max'],
                    # Defensive check: min/max should always be set when count > 0
                    'range': stats['max'] - stats['min'] if stats['max'] is not None and stats['min'] is not None else 0
                }
        
        analysis['drivability_p1_distribution'] = dict(driv_p1_counter)
        analysis['vehicle_data_summary'] = {
            'total_vehicles': len(all_vehicles),
            'vehicle_names': sorted(list(all_vehicles)),
            'rows_with_vehicle_data': vehicle_data_rows
        }
        analysis['per_vehicle_statistics'] = per_vehicle_stats
        
        # Data quality metrics
        missing_driv = driv_p1_counter.get('Unspecified', 0)
        analysis['data_quality'] = {
            'completeness_drivability': f"{((len(self.data_rows) - missing_driv) / len(self.data_rows) * 100):.1f}%",
            'completeness_vehicle_data': f"{(vehicle_data_rows / len(self.data_rows) * 100):.1f}%",
            'missing_drivability_p1': missing_driv,
            'rows_without_vehicle_data': len(self.data_rows) - vehicle_data_rows
        }
        
        # Generate insights
        self._generate_insights(analysis)
        
        return analysis
    
    def _generate_insights(self, analysis: Dict):
        """Generate insights from the analysis"""
        insights = []
        
        total = analysis['total_rows']
        
        # Drivability insights
        driv_dist = analysis['drivability_p1_distribution']
        red_driv = driv_dist.get('Red', 0)
        yellow_driv = driv_dist.get('Yellow', 0)
        green_driv = driv_dist.get('Green', 0)
        unspecified_driv = driv_dist.get('Unspecified', 0)
        
        if red_driv > 0:
            insights.append(f"⚠️  {red_driv} use cases ({red_driv/total*100:.1f}%) have RED drivability status - these need immediate attention")
        
        if yellow_driv > 0:
            insights.append(f"⚡ {yellow_driv} use cases ({yellow_driv/total*100:.1f}%) have YELLOW drivability status - these require improvement")
        
        if green_driv > 0:
            insights.append(f"✅ {green_driv} use cases ({green_driv/total*100:.1f}%) have GREEN drivability status - performing well")
        
        if unspecified_driv == total:
            insights.append(f"ℹ️  No P1 status data found - this is expected if users haven't pasted test data yet or if the VBA macros haven't been run")
        elif unspecified_driv > total * 0.5:
            insights.append(f"📊 Data Quality Issue: {unspecified_driv} rows ({unspecified_driv/total*100:.1f}%) missing drivability P1 status")
        
        # Overall performance
        if green_driv > 0:
            if green_driv >= total * 0.8:
                insights.append(f"🎯 Excellent overall performance! {green_driv/total*100:.1f}% of use cases are GREEN")
            elif red_driv >= total * 0.3:
                insights.append(f"🔴 Critical: {red_driv/total*100:.1f}% of use cases are RED - significant improvements needed")
        
        # Vehicle data insights
        vehicle_summary = analysis['vehicle_data_summary']
        if vehicle_summary['total_vehicles'] > 0:
            insights.append(f"🚗 Found test data for {vehicle_summary['total_vehicles']} vehicle(s): {', '.join(vehicle_summary['vehicle_names'][:3])}" + 
                          (f" and {vehicle_summary['total_vehicles'] - 3} more" if vehicle_summary['total_vehicles'] > 3 else ""))
            
            if vehicle_summary['rows_with_vehicle_data'] < total * 0.5:
                insights.append(f"📊 Only {vehicle_summary['rows_with_vehicle_data']}/{total} use cases have vehicle test data")
        else:
            insights.append(f"ℹ️  No vehicle test data detected - users should paste test data into vehicle columns (typically starting from column L)")
        
        analysis['insights'] = insights
    
    def print_analysis(self, analysis: Dict):
        """Print analysis results in a readable format"""
        print("\n" + "="*80)
        print("EVALUATION DATA ANALYSIS REPORT")
        print("="*80)
        
        print(f"\n📋 Total Use Cases Analyzed: {analysis['total_rows']}")
        
        print("\n" + "-"*80)
        print("DRIVABILITY P1 STATUS DISTRIBUTION")
        print("-"*80)
        for status, count in sorted(analysis['drivability_p1_distribution'].items()):
            percentage = (count / analysis['total_rows']) * 100
            print(f"  {status:15s}: {count:4d} ({percentage:5.1f}%)")
        
        # Vehicle data summary
        vehicle_summary = analysis.get('vehicle_data_summary', {})
        if vehicle_summary.get('total_vehicles', 0) > 0:
            print("\n" + "-"*80)
            print("VEHICLE TEST DATA SUMMARY")
            print("-"*80)
            print(f"  Total Vehicles: {vehicle_summary['total_vehicles']}")
            print(f"  Rows with Data: {vehicle_summary['rows_with_vehicle_data']} / {analysis['total_rows']}")
            if vehicle_summary['vehicle_names']:
                print(f"\n  Vehicles:")
                for vehicle in vehicle_summary['vehicle_names']:
                    print(f"    - {vehicle}")
        
        # Per-vehicle statistics
        per_vehicle = analysis.get('per_vehicle_statistics', {})
        if per_vehicle:
            print("\n" + "-"*80)
            print("PER-VEHICLE STATISTICS (Numeric Data)")
            print("-"*80)
            for vehicle, stats in sorted(per_vehicle.items()):
                print(f"\n  {vehicle}:")
                print(f"    Data Points: {stats['data_points']}")
                print(f"    Average:     {stats['average']:.2f}")
                print(f"    Min:         {stats['min']:.2f}")
                print(f"    Max:         {stats['max']:.2f}")
                print(f"    Range:       {stats['range']:.2f}")
        
        print("\n" + "-"*80)
        print("DATA QUALITY METRICS")
        print("-"*80)
        dq = analysis['data_quality']
        print(f"  Drivability P1 Completeness: {dq['completeness_drivability']}")
        print(f"  Vehicle Data Completeness:   {dq['completeness_vehicle_data']}")
        
        if analysis['insights']:
            print("\n" + "-"*80)
            print("KEY INSIGHTS")
            print("-"*80)
            for insight in analysis['insights']:
                print(f"  {insight}")
        
        print("\n" + "="*80)
    
    def export_to_json(self, output_path: str, analysis: Dict):
        """Export analysis results to JSON file"""
        with open(output_path, 'w') as f:
            json.dump(analysis, f, indent=2)
        print(f"\n✅ Analysis exported to: {output_path}")
    
    def run(self, output_json: Optional[str] = None):
        """Run the complete analysis"""
        print(f"📂 Loading Excel file: {self.excel_path}")
        
        if not self.find_evaluation_sheet():
            print("❌ Error: Could not find Evaluation Sheet")
            return None
        
        print(f"✅ Found Evaluation Sheet: {self.eval_sheet.title}")
        
        if not self.detect_header():
            print("❌ Error: Could not detect header row with 'USE CASE'")
            return None
        
        print(f"✅ Header detected at row {self.header_row}, column {self.use_case_col}")
        
        print("📊 Extracting data rows...")
        self.extract_data_rows()
        
        if not self.data_rows:
            print("❌ Error: No data rows found")
            return None
        
        print(f"✅ Extracted {len(self.data_rows)} data rows")
        
        print("🔍 Analyzing data...")
        analysis = self.analyze_data()
        
        self.print_analysis(analysis)
        
        if output_json:
            self.export_to_json(output_json, analysis)
        
        return analysis


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Analyze evaluation metric data from WL EVO AVLDrive Heatmap Tool',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s WL_EVO_AVLDrive_Heatmap_Tool.xlsm
  %(prog)s WL_EVO_AVLDrive_Heatmap_Tool.xlsm --output analysis_results.json
        """
    )
    
    parser.add_argument(
        'excel_file',
        help='Path to the Excel file (WL_EVO_AVLDrive_Heatmap_Tool.xlsm)'
    )
    
    parser.add_argument(
        '-o', '--output',
        help='Output JSON file for analysis results',
        default=None
    )
    
    args = parser.parse_args()
    
    try:
        analyzer = EvaluationDataAnalyzer(args.excel_file)
        analyzer.run(output_json=args.output)
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
